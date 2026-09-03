"""Exports: CSV, XLSX, PDF, the printable rescan checklist, and the flagged-page ZIP.

Every export in this module is a **clinical record extract**, and that dictates three things:

* Diagnosis text is written verbatim. The CSV and the XLSX carry the full transcription, never a
  summary and never a silent truncation. The PDF, which has a fixed page width, shortens a cell
  only when it must and always marks it with a trailing "…" plus a note telling the reader where
  the complete text lives.
* Nothing is presented as fact that a human has not checked. Every rendered artefact repeats the
  line "AI findings are not clinically confirmed unless marked reviewer-confirmed.", and every row
  carries ``ai_vs_reviewed``.
* Originals and annotated copies are kept apart. In the ZIP they live in separate top-level folders
  with a README that says which is which; an annotated PNG additionally has that fact burnt into
  the image itself (see ``app.services.annotate``).

All four writers take the rows produced by :func:`app.services.query.page_rows`, so the spreadsheet,
the PDF and the dashboard can never disagree about what the filter selected.
"""

from __future__ import annotations

import csv
import io
import os
import re
import zipfile
from collections import OrderedDict
from datetime import datetime, timezone
from typing import Any, Iterable, Sequence

from sqlalchemy.orm import Session

from app.services.annotate import ALL_LAYERS, annotate_page
from app.services.query import (
    EXPORT_COLUMNS,
    EXPORT_HEADERS,
    REVIEWER_CONFIRMED,
    REVIEWER_CORRECTED,
    PageFilters,
)

# --------------------------------------------------------------------- notices

DISCLAIMER = "AI findings are not clinically confirmed unless marked reviewer-confirmed."

TRUNCATION_NOTE = (
    "Cells ending in a single ellipsis character have been shortened to fit this page. "
    "The CSV and XLSX exports of the same query contain the complete text."
)

_ELLIPSIS = "…"

#: Excel refuses to store more than this in one cell. We never drop text silently; if a diagnosis
#: is somehow longer than this it is marked and the CSV is named as the complete copy.
_XLSX_CELL_LIMIT = 32_000


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _stamp() -> str:
    return _now().strftime("%Y-%m-%d %H:%M:%S UTC")


def _as_rows(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """Materialise once — every writer walks the rows more than once."""
    return rows if isinstance(rows, list) else list(rows)


def _cell(row: dict[str, Any], key: str) -> str:
    value = row.get(key, "")
    if value is None:
        return ""
    return str(value)


def _params_pairs(params: PageFilters | dict[str, Any] | None) -> list[tuple[str, str]]:
    """Normalise whatever the route had into (label, value) pairs for the parameters block."""
    if params is None:
        return [("Filters", "not recorded")]
    if isinstance(params, PageFilters):
        return params.describe()
    pairs = [(str(k), "; ".join(str(x) for x in v) if isinstance(v, (list, tuple)) else str(v))
             for k, v in params.items() if v not in (None, "", [], ())]
    return pairs or [("Filters", "none — all active page versions")]


# ------------------------------------------------------------------------ CSV


def export_csv(rows: Iterable[dict[str, Any]]) -> bytes:
    """The complete extract, one row per active page version.

    Encoded UTF-8 **with a BOM**: without it Excel on Windows mis-reads Devanagari and any
    non-ASCII patient reference, and a mangled clinical extract is worse than no extract.
    Line endings are CRLF for the same reason.
    """
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([EXPORT_HEADERS[c] for c in EXPORT_COLUMNS])
    for row in rows:
        # Full text, verbatim, never truncated. Newlines inside a diagnosis are preserved by the
        # csv module's own quoting.
        writer.writerow([_cell(row, c) for c in EXPORT_COLUMNS])
    return buf.getvalue().encode("utf-8-sig")


# ----------------------------------------------------------------------- XLSX


def export_xlsx(
    rows: Iterable[dict[str, Any]],
    params: PageFilters | dict[str, Any] | None = None,
    summary: dict[str, Any] | None = None,
) -> bytes:
    """A two-sheet workbook: "Report parameters" first, then "Pages".

    The parameters sheet leads deliberately. A spreadsheet of 4,000 rows with no record of which
    filter produced it invites someone to read it as "the whole batch", and that is exactly the
    misreading that gets a missing page signed off.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = _as_rows(rows)
    wb = Workbook()

    # ------------------------------------------------ sheet 1: parameters
    meta = wb.active
    meta.title = "Report parameters"
    meta.column_dimensions["A"].width = 34
    meta.column_dimensions["B"].width = 80

    bold = Font(bold=True)
    heading = Font(bold=True, size=13)

    meta.append(["Scan quality report"])
    meta["A1"].font = heading
    meta.append(["Generated (UTC)", _stamp()])
    meta.append(["Rows in this report", len(rows)])
    meta.append([])
    meta.append(["Filters applied"])
    meta.cell(row=meta.max_row, column=1).font = bold
    for label, value in _params_pairs(params):
        meta.append([label, value])

    if summary:
        meta.append([])
        meta.append(["Totals for this selection"])
        meta.cell(row=meta.max_row, column=1).font = bold
        for label, value in _summary_pairs(summary):
            meta.append([label, value])

    meta.append([])
    meta.append(["Important", DISCLAIMER])
    meta.cell(row=meta.max_row, column=1).font = bold
    meta.append([
        "Note",
        "'unchecked', 'blank' and 'failed' pages are NOT acceptable pages. A handwriting status of "
        "'failed' or 'unconfigured' means NOT CHECKED, not 'no handwriting found'.",
    ])
    meta.append([
        "Note",
        "Diagnosis text in this workbook is reproduced verbatim from the record. Column "
        "'AI vs reviewed' states whether a human confirmed or corrected it.",
    ])
    for row_idx in range(1, meta.max_row + 1):
        meta.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # ----------------------------------------------------- sheet 2: pages
    ws = wb.create_sheet("Pages")
    headers = [EXPORT_HEADERS[c] for c in EXPORT_COLUMNS]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="DDE5EE")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    widths = [len(h) + 2 for h in headers]
    for row in rows:
        values: list[Any] = []
        for i, key in enumerate(EXPORT_COLUMNS):
            value = row.get(key, "")
            if isinstance(value, str) and len(value) > _XLSX_CELL_LIMIT:
                # Only ever reached by a pathological transcription; marked, never silent.
                value = value[: _XLSX_CELL_LIMIT - 1] + _ELLIPSIS
            elif value is None:
                value = ""
            values.append(value)
            # Auto-ish widths: sample the longest line of each cell, capped so one long diagnosis
            # does not produce a 900-character-wide column.
            text = str(value)
            longest = max((len(part) for part in text.split("\n")), default=0)
            if longest + 2 > widths[i]:
                widths[i] = min(60, longest + 2)
        ws.append(values)

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(8, width)

    ws.freeze_panes = "A2"  # header stays visible while scrolling a long batch
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, ws.max_row)}"

    # Wrap the two long free-text columns so the full value is readable in place.
    for key in ("diagnosis_text_raw", "diagnosis_text_reviewed", "reviewer_comment"):
        col = get_column_letter(EXPORT_COLUMNS.index(key) + 1)
        for cell in ws[col][1:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _summary_pairs(summary: dict[str, Any]) -> list[tuple[str, str]]:
    """Flatten the dashboard payload into label/value lines, keeping its own wording."""
    totals = summary.get("totals", summary) or {}
    overlaps = summary.get("overlaps", {}) or {}
    pairs: list[tuple[str, str]] = [
        ("Files", str(totals.get("files", ""))),
        ("Active pages", str(totals.get("pages_active", ""))),
        ("Awaiting review", str(totals.get("awaiting_review", ""))),
    ]
    for group in ("quality", "handwriting", "diagnosis"):
        block = totals.get(group) or {}
        if block:
            pairs.append((group.capitalize(), ", ".join(f"{k}: {v}" for k, v in block.items())))
    if overlaps:
        pairs.append(
            (
                "Overlap",
                f"defect and handwriting: {overlaps.get('defect_and_handwriting', 0)}, "
                f"defect only: {overlaps.get('defect_only', 0)}, "
                f"handwriting only: {overlaps.get('handwriting_only', 0)}",
            )
        )
    return pairs


# ------------------------------------------------------------------ PDF setup


def _register_font() -> tuple[str, str, str | None]:
    """Register a Unicode TTF if one is installed; otherwise fall back to Helvetica.

    Returns ``(regular, bold, warning)``. Helvetica cannot render Devanagari at all, so when no
    suitable font is present the PDF says so on its own front matter rather than printing boxes and
    letting a reader assume the record was empty.
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        ("NotoDevanagari", "/usr/share/fonts/truetype/noto/NotoSansDevanagari-Regular.ttf",
         "/usr/share/fonts/truetype/noto/NotoSansDevanagari-Bold.ttf"),
        ("Lohit", "/usr/share/fonts/truetype/lohit-devanagari/Lohit-Devanagari.ttf", None),
        ("DejaVu", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("FreeSans", "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
         "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf"),
    ]
    for name, regular, bold in candidates:
        if not os.path.exists(regular):
            continue
        try:
            pdfmetrics.registerFont(TTFont(name, regular))
            bold_name = name
            if bold and os.path.exists(bold):
                bold_name = f"{name}-Bold"
                pdfmetrics.registerFont(TTFont(bold_name, bold))
            warning = None
            if "Devanagari" not in regular and "Lohit" not in regular:
                warning = (
                    "This PDF is set in a font without Devanagari coverage. Hindi text may not "
                    "render; use the CSV or XLSX export for the authoritative text."
                )
            return name, bold_name, warning
        except Exception:
            continue
    return (
        "Helvetica",
        "Helvetica-Bold",
        "This PDF is set in Helvetica, which cannot render Devanagari. Non-Latin text may appear "
        "blank or as boxes; use the CSV or XLSX export for the authoritative text.",
    )


def _make_canvas(font: str, footer_lines: Sequence[str]):
    """A canvas class that stamps "Page X of Y" and the standing notices on every page.

    ReportLab does not know the page count until the document is finished, so pages are held and
    the footer is written on a second pass. The disclaimer goes on *every* page: a report read one
    page at a time must never lose the caveat that sat on page one.
    """
    from reportlab.pdfgen import canvas as _canvas

    class NumberedCanvas(_canvas.Canvas):
        def __init__(self, *args: Any, **kwargs: Any) -> None:
            super().__init__(*args, **kwargs)
            self._saved: list[dict[str, Any]] = []

        def showPage(self) -> None:  # noqa: N802 - reportlab API
            self._saved.append(dict(self.__dict__))
            self._startPage()

        def save(self) -> None:
            total = len(self._saved)
            for state in self._saved:
                self.__dict__.update(state)
                self._draw_footer(total)
                super().showPage()
            super().save()

        def _draw_footer(self, total: int) -> None:
            width, _ = self._pagesize
            self.setFont(font, 7)
            self.setFillGray(0.35)
            y = 16
            for line in reversed(footer_lines):
                self.drawString(24, y, line)
                y += 9
            self.setFont(font, 7.5)
            self.setFillGray(0.2)
            self.drawRightString(width - 24, 16, f"Page {self._pageNumber} of {total}")

    return NumberedCanvas


def _shorten(text: str, limit: int) -> str:
    """Shorten for layout ONLY, and mark it. Never used by the CSV or XLSX writers."""
    text = (text or "").replace("\r", " ")
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + _ELLIPSIS


# ------------------------------------------------------------------- PDF: pages

#: Columns the landscape PDF can carry legibly, with relative widths. The remaining columns are in
#: the CSV/XLSX; the PDF says so on its front matter rather than pretending it is complete.
#: The short heading is used where the full one would wrap mid-word in a narrow column — an
#: abbreviation only, never a different meaning.
_PDF_COLUMNS: tuple[tuple[str, float, int, str], ...] = (
    # (column key, relative width, character budget before layout shortening, heading)
    ("batch", 7.0, 40, "Batch"),
    ("patient_ref", 6.5, 32, "Patient ref"),
    ("encounter_ref", 6.5, 32, "Encounter ref"),
    ("filename", 11.0, 60, "File"),
    ("page_no", 2.6, 8, "Pg"),
    ("version", 2.8, 6, "Ver"),
    ("scan_status", 6.0, 16, "Scan status"),
    ("defect_codes", 10.0, 60, "Defect codes"),
    ("defect_severities", 7.0, 40, "Defect severities"),
    ("handwriting_status", 8.0, 20, "Handwriting"),
    ("diagnosis_status", 8.0, 28, "Diagnosis status"),
    ("diagnosis_text_raw", 15.6, 300, "Diagnosis text (AI, raw)"),
    ("ai_vs_reviewed", 8.5, 22, "AI vs reviewed"),
)


def export_pdf(
    rows: Iterable[dict[str, Any]],
    summary: dict[str, Any] | None = None,
    params: PageFilters | dict[str, Any] | None = None,
    title: str = "Scan quality report",
) -> bytes:
    """Landscape A4 report: totals block, repeating table header, page numbers, footer notice.

    The totals block is the dashboard's own payload, so a printed report and the screen a colleague
    is looking at cannot show different numbers.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        BaseDocTemplate,
        Frame,
        KeepTogether,
        PageTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    rows = _as_rows(rows)
    font, font_bold, font_warning = _register_font()

    page_size = landscape(A4)
    margin = 12 * mm
    avail_w = page_size[0] - 2 * margin

    body = ParagraphStyle("body", fontName=font, fontSize=7.2, leading=8.6, alignment=TA_LEFT)
    cell = ParagraphStyle("cell", parent=body, fontSize=6.6, leading=7.8)
    cell_head = ParagraphStyle("cellhead", parent=cell, fontName=font_bold, textColor=colors.white)
    h1 = ParagraphStyle("h1", fontName=font_bold, fontSize=14, leading=17, spaceAfter=4)
    h2 = ParagraphStyle("h2", fontName=font_bold, fontSize=9, leading=12, spaceBefore=5, spaceAfter=2)
    grey = colors.HexColor("#555555")
    note = ParagraphStyle("note", parent=body, fontSize=6.8, leading=8.4, textColor=grey)

    story: list[Any] = [Paragraph(title, h1)]
    story.append(Paragraph(f"Generated {_stamp()} &nbsp;|&nbsp; {len(rows)} active page versions", body))

    # ---------------------------------------------------------- parameters
    story.append(Paragraph("Filters applied", h2))
    param_rows = [[Paragraph(f"<b>{_esc(label)}</b>", cell), Paragraph(_esc(value), cell)]
                  for label, value in _params_pairs(params)]
    table = Table(param_rows, colWidths=[avail_w * 0.18, avail_w * 0.82], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(table)

    # ------------------------------------------------------------- totals
    if summary:
        story.append(Paragraph("Totals for this selection", h2))
        story.append(_summary_table(summary, avail_w, cell, font_bold, colors, Table, TableStyle))
        story.append(
            Paragraph(
                "'blank', 'failed' and 'unchecked' are separate outcomes and are not included in "
                "'acceptable'. Handwriting is counted on its own axis and is never treated as a "
                "scan defect; the overlap line shows how the two sets intersect.",
                note,
            )
        )

    # -------------------------------------------------------------- table
    story.append(Paragraph("Pages", h2))
    total_weight = sum(w for _, w, _, _ in _PDF_COLUMNS)
    col_widths = [avail_w * (w / total_weight) for _, w, _, _ in _PDF_COLUMNS]

    data: list[list[Any]] = [
        [Paragraph(_esc(heading), cell_head) for _, _, _, heading in _PDF_COLUMNS]
    ]
    shortened = False
    for row in rows:
        line: list[Any] = []
        for key, _, budget, _heading in _PDF_COLUMNS:
            text = _cell(row, key)
            clipped = _shorten(text, budget)
            if clipped != text:
                shortened = True
            line.append(Paragraph(_esc(clipped).replace("\n", "<br/>"), cell))
        data.append(line)

    if len(data) == 1:
        data.append([Paragraph("No pages matched these filters.", cell)] + [""] * (len(_PDF_COLUMNS) - 1))

    table = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F4356")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9AA6B2")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    for i, row in enumerate(rows, start=1):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F4F6F8")))
        # A reviewer-confirmed row is tinted so the eye can find the checked findings on a printout.
        if row.get("ai_vs_reviewed") in (REVIEWER_CONFIRMED, REVIEWER_CORRECTED):
            style.append(("BACKGROUND", (len(_PDF_COLUMNS) - 1, i), (-1, i), colors.HexColor("#E4F1E4")))
    table.setStyle(TableStyle(style))
    story.append(table)

    tail = [
        Paragraph(
            "Columns not shown here (printed label, handwriting categories, diagnosis qualifier, "
            "reviewer-corrected text, reviewer, comment) are present in the CSV and XLSX exports "
            "of this same query.",
            note,
        )
    ]
    if shortened:
        tail.append(Paragraph(TRUNCATION_NOTE, note))
    if font_warning:
        tail.append(Paragraph(font_warning, note))
    story.append(Spacer(1, 4))
    story.append(KeepTogether(tail))

    footer = [DISCLAIMER, f"{title} — generated {_stamp()}"]
    if shortened:
        footer.insert(1, f"Cells ending in {_ELLIPSIS} are shortened for layout; see the CSV/XLSX export.")

    out = io.BytesIO()
    doc = BaseDocTemplate(
        out,
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin + 14,  # room for the footer block
        title=title,
        author="OPD Scan QC",
    )
    frame = Frame(margin, doc.bottomMargin, avail_w, page_size[1] - margin - doc.bottomMargin, id="body")
    doc.addPageTemplates([PageTemplate(id="all", frames=[frame])])
    doc.build(story, canvasmaker=_make_canvas(font, footer))
    return out.getvalue()


def _summary_table(summary, avail_w, cell_style, font_bold, colors, Table, TableStyle):  # noqa: ANN001
    """The dashboard totals as a compact grid."""
    from reportlab.platypus import Paragraph

    pairs = _summary_pairs(summary)
    data = [[Paragraph(f"<b>{_esc(label)}</b>", cell_style), Paragraph(_esc(value), cell_style)]
            for label, value in pairs]
    table = Table(data, colWidths=[avail_w * 0.18, avail_w * 0.82], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#C8D0D8")),
            ]
        )
    )
    return table


def _esc(value: Any) -> str:
    """Escape for reportlab's mini-HTML. Patient text can legitimately contain & and <."""
    return (
        str(value if value is not None else "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


# -------------------------------------------------------- PDF: rescan checklist


def rescan_checklist_pdf(
    rows: Iterable[dict[str, Any]],
    params: PageFilters | dict[str, Any] | None = None,
) -> bytes:
    """A printable work list for the scanning desk, grouped by batch then by file.

    Portrait, wide tick column, one line per page to re-scan with the reason and severity. Designed
    to be carried to the scanner and marked off with a pen, so it repeats the identifying details
    (file name, page number, printed label) rather than relying on an on-screen selection.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        BaseDocTemplate,
        Frame,
        KeepTogether,
        PageTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    rows = _as_rows(rows)
    font, font_bold, font_warning = _register_font()

    page_size = A4
    margin = 14 * mm
    avail_w = page_size[0] - 2 * margin

    body = ParagraphStyle("body", fontName=font, fontSize=8.5, leading=10.5, alignment=TA_LEFT)
    cell = ParagraphStyle("cell", parent=body, fontSize=7.6, leading=9.2)
    cell_head = ParagraphStyle("ch", parent=cell, fontName=font_bold)
    h1 = ParagraphStyle("h1", fontName=font_bold, fontSize=15, leading=18, spaceAfter=3)
    h2 = ParagraphStyle("h2", fontName=font_bold, fontSize=10.5, leading=13, spaceBefore=8, spaceAfter=2)
    h3 = ParagraphStyle("h3", fontName=font_bold, fontSize=9, leading=11, spaceBefore=4, spaceAfter=1)
    note = ParagraphStyle("note", parent=body, fontSize=7.4, leading=9, textColor=colors.HexColor("#555555"))

    # Group: batch → file → pages, preserving the incoming (already sorted) order.
    grouped: "OrderedDict[str, OrderedDict[str, list[dict[str, Any]]]]" = OrderedDict()
    for row in rows:
        batch = _cell(row, "batch") or "(no batch)"
        filename = _cell(row, "filename") or "(no file)"
        grouped.setdefault(batch, OrderedDict()).setdefault(filename, []).append(row)

    story: list[Any] = [
        Paragraph("Rescan work list", h1),
        Paragraph(
            f"Generated {_stamp()} &nbsp;|&nbsp; {len(rows)} page(s) to re-scan across "
            f"{len(grouped)} batch(es)",
            body,
        ),
        Paragraph(
            "Re-scan each ticked page and upload it against the same file. The system will create a "
            "new version of that page; the earlier scan is kept but stops being counted.",
            note,
        ),
    ]
    if params is not None:
        story.append(
            Paragraph(
                "Selection: " + "; ".join(f"{k}: {v}" for k, v in _params_pairs(params)), note
            )
        )

    col_widths = [
        avail_w * 0.08,  # tick
        avail_w * 0.07,  # page
        avail_w * 0.09,  # printed label
        avail_w * 0.06,  # version
        avail_w * 0.10,  # status
        avail_w * 0.42,  # reason
        avail_w * 0.18,  # severity
    ]
    headers = ["Done", "Page", "Printed", "Ver", "Status", "Reason", "Severity"]

    for batch, files in grouped.items():
        story.append(Paragraph(f"Batch: {_esc(batch)}", h2))
        for filename, page_rows_ in files.items():
            first = page_rows_[0]
            refs = " / ".join(
                x for x in (_cell(first, "patient_ref"), _cell(first, "encounter_ref")) if x
            )
            block: list[Any] = [
                Paragraph(_esc(filename) + (f" &nbsp;<font size=7>[{_esc(refs)}]</font>" if refs else ""), h3)
            ]
            data: list[list[Any]] = [[Paragraph(_esc(h), cell_head) for h in headers]]
            for row in page_rows_:
                reason = _reason_text(row)
                data.append(
                    [
                        Paragraph("", cell),  # the tick box is drawn by the table style
                        Paragraph(_esc(_cell(row, "page_no")), cell),
                        Paragraph(_esc(_cell(row, "printed_label")), cell),
                        Paragraph(_esc(_cell(row, "version")), cell),
                        Paragraph(_esc(_cell(row, "scan_status")), cell),
                        Paragraph(_esc(_shorten(reason, 220)), cell),
                        Paragraph(_esc(_cell(row, "defect_severities") or "-"), cell),
                    ]
                )
            table = Table(data, colWidths=col_widths, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EDF2")),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#8E9AA6")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                        # The tick column: a heavy empty cell that reads as a box to mark.
                        ("BOX", (0, 1), (0, -1), 0.9, colors.black),
                        ("INNERGRID", (0, 1), (0, -1), 0.9, colors.black),
                    ]
                )
            )
            block.append(table)
            block.append(Spacer(1, 3))
            story.append(KeepTogether(block))

    if not grouped:
        story.append(Paragraph("No pages currently need a re-scan for this selection.", body))
    if font_warning:
        story.append(Spacer(1, 4))
        story.append(Paragraph(font_warning, note))

    footer = [DISCLAIMER, f"Rescan work list — generated {_stamp()}"]

    out = io.BytesIO()
    doc = BaseDocTemplate(
        out,
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin + 14,
        title="Rescan work list",
        author="OPD Scan QC",
    )
    frame = Frame(margin, doc.bottomMargin, avail_w, page_size[1] - margin - doc.bottomMargin, id="body")
    doc.addPageTemplates([PageTemplate(id="all", frames=[frame])])
    doc.build(story, canvasmaker=_make_canvas(font, footer))
    return out.getvalue()


def _reason_text(row: dict[str, Any]) -> str:
    """Why this page is on the list, in words a scanning clerk can act on."""
    findings = row.get("_quality_findings") or []
    parts: list[str] = []
    seen: set[str] = set()
    for finding in findings:
        code = str(finding.get("defect_code", ""))
        if code in seen:
            continue
        seen.add(code)
        label = finding.get("label") or code
        parts.append(f"{label} ({finding.get('severity', '?')})")
    if not parts and row.get("defect_codes"):
        parts.append(str(row["defect_codes"]))
    if row.get("ai_vs_reviewed") in (REVIEWER_CONFIRMED, REVIEWER_CORRECTED):
        parts.append("reviewer confirmed")
    if not parts:
        parts.append("Requested by reviewer")
    return "; ".join(parts)


# ------------------------------------------------------------------- ZIP


_SAFE = re.compile(r"[^A-Za-z0-9._-]+")


def _safe(part: str, fallback: str = "unnamed") -> str:
    """A path component that is safe on every filesystem a ZIP might be unpacked on."""
    cleaned = _SAFE.sub("_", (part or "").strip()).strip("._")
    return (cleaned or fallback)[:80]


def flagged_zip(
    db: Session,
    rows: Iterable[dict[str, Any]],
    annotated: bool = False,
    params: PageFilters | dict[str, Any] | None = None,
) -> bytes:
    """A ZIP of the flagged page images plus a manifest and a README.

    Layout — originals and annotated copies are **never** mixed::

        originals/<batch>/<file>/p0007_v1.png     the untouched stored render
        annotated/<batch>/<file>/p0007_v1.png     the same page with overlays burnt in
        manifest.csv                              every export column plus the ZIP paths
        README.txt                                which folder is which, and the AI caveat

    ``annotated=False`` produces no ``annotated/`` folder at all, so a reader can never mistake an
    overlay render for the scan of record. Every annotated PNG also carries the "ANNOTATED — not
    the original scan" caption inside the image.
    """
    from app.core.storage import get_storage
    from app.models.core import PageVersion
    from app.processing.ingest import bytes_to_image, encode_png

    rows = _as_rows(rows)
    storage = get_storage()
    out = io.BytesIO()

    manifest_buf = io.StringIO(newline="")
    manifest = csv.writer(manifest_buf, lineterminator="\r\n")
    manifest.writerow(
        [EXPORT_HEADERS[c] for c in EXPORT_COLUMNS] + ["original_path", "annotated_path", "note"]
    )

    written = 0
    annotated_written = 0
    problems: list[str] = []

    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        used: set[str] = set()
        for row in rows:
            rel = _relative_name(row, used)
            original_path = f"originals/{rel}"
            annotated_path = ""
            note = ""

            key = row.get("_storage_key_render")
            pv = None
            if not key and row.get("_page_version_id"):
                # Rows that did not come from page_rows() — fall back to the database rather than
                # dropping the page from the ZIP without saying so.
                pv = db.get(PageVersion, row["_page_version_id"])
                key = pv.storage_key_render if pv else None

            data: bytes | None = None
            if key:
                try:
                    data = storage.get_bytes(key)
                except Exception as exc:
                    note = f"image unavailable: {type(exc).__name__}"
            else:
                note = "no stored render for this page version"

            if data:
                zf.writestr(original_path, data)
                written += 1
            else:
                original_path = ""
                problems.append(f"{rel}: {note}")

            if annotated and data is not None:
                try:
                    image = bytes_to_image(data)
                    if image is None:
                        raise ValueError("render could not be decoded")
                    overlay = annotate_page(
                        image,
                        row.get("_quality_findings") or [],
                        row.get("_handwriting_regions") or [],
                        row.get("_diagnosis_regions") or [],
                        set(ALL_LAYERS),
                    )
                    annotated_path = f"annotated/{rel}"
                    zf.writestr(annotated_path, encode_png(overlay))
                    annotated_written += 1
                except Exception as exc:
                    annotated_path = ""
                    note = (note + "; " if note else "") + f"annotation failed: {type(exc).__name__}"
                    problems.append(f"{rel}: annotation failed ({type(exc).__name__})")

            manifest.writerow(
                [_cell(row, c) for c in EXPORT_COLUMNS] + [original_path, annotated_path, note]
            )

        zf.writestr("manifest.csv", manifest_buf.getvalue().encode("utf-8-sig"))
        zf.writestr(
            "README.txt",
            _readme(
                total=len(rows),
                originals=written,
                annotated_count=annotated_written if annotated else None,
                params=params,
                problems=problems,
            ).encode("utf-8"),
        )

    return out.getvalue()


def _relative_name(row: dict[str, Any], used: set[str]) -> str:
    """``<batch>/<file>/p0007_v1.png``, unique within the archive."""
    batch = _safe(_cell(row, "batch"), "no_batch")
    filename = _safe(os.path.splitext(_cell(row, "filename"))[0], "no_file")
    try:
        page = int(row.get("page_no") or 0)
    except (TypeError, ValueError):
        page = 0
    version = row.get("version") or 1
    base = f"{batch}/{filename}/p{page:04d}_v{version}"
    candidate = f"{base}.png"
    n = 2
    while candidate in used:
        candidate = f"{base}__{n}.png"
        n += 1
    used.add(candidate)
    return candidate


def _readme(
    total: int,
    originals: int,
    annotated_count: int | None,
    params: PageFilters | dict[str, Any] | None,
    problems: Sequence[str],
) -> str:
    lines = [
        "FLAGGED PAGES EXPORT",
        "=" * 64,
        f"Generated: {_stamp()}",
        f"Pages selected: {total}",
        "",
        "WHAT IS IN THIS ARCHIVE",
        "-" * 64,
        f"originals/   {originals} file(s). The stored page render, byte-for-byte as the system",
        "             holds it. This is the scan of record.",
    ]
    if annotated_count is None:
        lines += [
            "annotated/   NOT INCLUDED in this archive. Re-export with annotated=true if you need",
            "             the overlay renders.",
        ]
    else:
        lines += [
            f"annotated/   {annotated_count} file(s). A COPY of the same page with detection overlays",
            "             drawn on top. These are NOT scans of the document and must never be filed",
            "             as such. Each one carries the caption 'ANNOTATED - not the original scan'",
            "             burnt into the image and a coloured frame around it.",
            "             Overlay colours: red = scan defects, blue = handwriting, green = diagnosis.",
        ]
    lines += [
        "manifest.csv One row per selected page, with every reporting column plus the paths above.",
        "",
        "HOW TO READ THE MANIFEST",
        "-" * 64,
        "'Scan status'  acceptable | review | rescan | blank | failed | unchecked.",
        "               blank, failed and unchecked are NOT acceptable pages.",
        "'Handwriting'  detected | none_detected | failed | unconfigured | pending.",
        "               'failed' and 'unconfigured' mean NOT CHECKED - not 'no handwriting'.",
        "'AI vs reviewed'",
        "               ai_only            - no human has reviewed this page's findings",
        "               reviewer_confirmed - a reviewer agreed with the AI output",
        "               reviewer_corrected - a reviewer changed the AI output",
        "",
        DISCLAIMER,
        "",
    ]
    if params is not None:
        lines += ["SELECTION", "-" * 64]
        lines += [f"{label}: {value}" for label, value in _params_pairs(params)]
        lines.append("")
    if problems:
        lines += [
            "PAGES WITH PROBLEMS",
            "-" * 64,
            "These pages are listed in the manifest but their image could not be included:",
        ]
        lines += [f"  - {p}" for p in problems[:200]]
        if len(problems) > 200:
            lines.append(f"  ... and {len(problems) - 200} more")
        lines.append("")
    return "\n".join(lines)
